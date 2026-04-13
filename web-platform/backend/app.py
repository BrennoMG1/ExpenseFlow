"""
ExpenseFlow — Backend FastAPI (v2 — Storage Bucket)
"""

import os
import io
import logging
from pathlib import Path

import httpx
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv

from orchestrator import OrchestratorClient

load_dotenv()

# ------------------------------------------------------------------ #
#  Logging                                                             #
# ------------------------------------------------------------------ #
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("expenseflow")

# ------------------------------------------------------------------ #
#  App                                                                 #
# ------------------------------------------------------------------ #
app = FastAPI(title="ExpenseFlow API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

orchestrator = OrchestratorClient()
FRONTEND_DIR = Path(__file__).parent.parent / "frontend"

# ------------------------------------------------------------------ #
#  Storage Bucket (produção — Cloud Robot)                            #
# ------------------------------------------------------------------ #
BUCKET_NAME      = os.getenv("BUCKET_NAME", "ExpenseFlow-bucket")
BUCKET_FILE      = os.getenv("BUCKET_FILE", "ExpenseFlow_Dados.xlsx")
ORCHESTRATOR_URL = os.getenv("ORCHESTRATOR_URL", "").rstrip("/")
FOLDER_ID = os.getenv("FOLDER_ID", "7706457")

# ------------------------------------------------------------------ #
#  Fallback: arquivo local (desenvolvimento)                          #
# ------------------------------------------------------------------ #
_excel_env = os.getenv("EXCEL_PATH")
if _excel_env:
    EXCEL_PATH = Path(_excel_env)
else:
    EXCEL_PATH = (
        Path(__file__).parent   # .../web-platform/backend
        .parent                  # .../web-platform
        .parent                  # .../ExpenseFlow  (raiz do projeto)
        / "ExpenseFlow"          # .../ExpenseFlow/ExpenseFlow (pasta UiPath)
        / "Dados"
        / "ExpenseFlow_Dados.xlsx"
    )

# ------------------------------------------------------------------ #
#  Contas                                                              #
# ------------------------------------------------------------------ #
def _carregar_contas() -> list:
    contas = []
    i = 1
    while True:
        label   = os.getenv(f"CONTA_{i}_LABEL", "").strip()
        conn_id = os.getenv(f"CONTA_{i}_ID",    "").strip()
        if not label or not conn_id:
            break
        contas.append({"label": label, "connectionId": conn_id})
        log.info(f"[config] Conta carregada: {label}")
        i += 1

    if not contas:
        log.warning(
            "[config] Nenhuma conta encontrada no .env! "
            "Configure CONTA_1_LABEL e CONTA_1_ID no arquivo .env"
        )

    contas.append({"label": "Todas as contas", "connectionId": ""})
    return contas


CONTAS = _carregar_contas()

# ------------------------------------------------------------------ #
#  Modelos                                                             #
# ------------------------------------------------------------------ #
class ProcessRequest(BaseModel):
    connection_id: str

# ------------------------------------------------------------------ #
#  Helper: baixar Excel do Storage Bucket                             #
# ------------------------------------------------------------------ #
async def _baixar_excel_do_bucket() -> bytes:
    """
    Baixa o arquivo Excel do Storage Bucket via API do Orchestrator.
    Fluxo:
      1. Autentica e busca o ID do bucket pelo nome
      2. Obtém uma URI de download temporária (válida 5 min)
      3. Faz o download dos bytes do arquivo
    """
    try:
        token = await orchestrator._get_token()
        headers = {
            "Authorization":              f"Bearer {token}",
            "X-UIPATH-OrganizationUnitId": FOLDER_ID,
        }

        async with httpx.AsyncClient(timeout=30) as client:

            # 1 — Buscar o bucket pelo nome
            resp_buckets = await client.get(
                f"{ORCHESTRATOR_URL}/odata/StorageBuckets",
                headers=headers,
                params={
                    "$filter": f"Name eq '{BUCKET_NAME}'",
                    "$select": "Id,Name",
                },
            )
            resp_buckets.raise_for_status()
            buckets = resp_buckets.json().get("value", [])

            if not buckets:
                raise HTTPException(
                    status_code=404,
                    detail=f"Bucket '{BUCKET_NAME}' não encontrado no Orchestrator."
                )

            bucket_id = buckets[0]["Id"]
            log.info(f"[bucket] Bucket '{BUCKET_NAME}' encontrado. ID: {bucket_id}")

            # 2 — Obter URI de leitura temporária
            resp_uri = await client.get(
                f"{ORCHESTRATOR_URL}/odata/StorageBuckets({bucket_id})"
                f"/UiPath.Server.Configuration.OData.GetReadUri",
                headers=headers,
                params={"path": BUCKET_FILE, "expirationInMinutes": 5},
            )
            resp_uri.raise_for_status()
            download_url = resp_uri.json().get("Uri") or resp_uri.json().get("value")

            if not download_url:
                raise HTTPException(
                    status_code=502,
                    detail="Orchestrator não retornou URI de download do bucket."
                )

            # 3 — Baixar os bytes do arquivo
            resp_file = await client.get(download_url)
            resp_file.raise_for_status()
            log.info(f"[bucket] '{BUCKET_FILE}' baixado com sucesso ({len(resp_file.content)} bytes)")
            return resp_file.content

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"Erro ao acessar o Storage Bucket: {str(e)}"
        )


async def _obter_excel_bytes() -> bytes:
    """
    Estratégia de obtenção do Excel:
      - Produção  → baixa do Storage Bucket (Cloud Robot)
      - Dev local → lê do caminho local (fallback)
    """
    # Tenta o bucket
    if ORCHESTRATOR_URL:
        try:
            return await _baixar_excel_do_bucket()
        except HTTPException as e:
            log.warning(f"[bucket] Falha ({e.detail}). Tentando arquivo local como fallback...")

    raise HTTPException(
        status_code=404,
        detail="Nenhum dado encontrado. Execute o robô primeiro."
    )


# ------------------------------------------------------------------ #
#  Endpoints                                                           #
# ------------------------------------------------------------------ #

@app.get("/api/contas")
async def get_contas():
    """Retorna as contas disponíveis para o dropdown do frontend."""
    return {"contas": CONTAS}


@app.post("/api/processar")
async def processar(request: ProcessRequest):
    conta = next((c for c in CONTAS if c["connectionId"] == request.connection_id), None)
    label = conta["label"] if conta else "todas as contas"

    try:
        job = await orchestrator.start_job(request.connection_id)
        log.info(f"[job] Iniciado para '{label}' → JobId: {job['Id']}")
        return {
            "success": True,
            "jobId":   str(job["Id"]),
            "label":   label,
            "message": f"Processamento iniciado para: {label}",
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao iniciar o processo: {str(e)}")


@app.get("/api/status/{job_id}")
async def get_status(job_id: str):
    try:
        return await orchestrator.get_job_status(job_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao consultar status: {str(e)}")


@app.get("/api/dados")
async def get_dados():
    """Lê os dados do Excel (bucket ou local) e retorna como JSON para a tabela."""
    try:
        import openpyxl
        excel_bytes = await _obter_excel_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(c is not None for c in row):
                rows.append([str(v) if v is not None else "" for v in row])
        return {"headers": headers, "rows": rows, "total": len(rows)}

    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Dependência 'openpyxl' não instalada. Execute: pip install openpyxl"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler dados: {str(e)}")


@app.get("/api/download")
async def download_excel():
    """Baixa o arquivo Excel diretamente do Storage Bucket."""
    excel_bytes = await _obter_excel_bytes()
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ExpenseFlow_Dados.xlsx"},
    )


@app.get("/api/health")
async def health():
    return {
        "status":      "ok",
        "service":     "ExpenseFlow",
        "version":     "2.0.0",
        "bucket_name": BUCKET_NAME,
        "bucket_file": BUCKET_FILE,
        "local_excel": str(EXCEL_PATH),
        "local_ok":    EXCEL_PATH.exists(),
        "contas":      len([c for c in CONTAS if c["connectionId"]]),
    }


@app.get("/")
async def serve_index():
    return FileResponse(FRONTEND_DIR / "index.html")


@app.get("/{filename:path}")
async def serve_static(filename: str):
    file_path = FRONTEND_DIR / filename
    if file_path.exists() and file_path.is_file():
        return FileResponse(file_path)
    raise HTTPException(status_code=404, detail="Arquivo não encontrado.")


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("app:app", host="0.0.0.0", port=port, reload=True)